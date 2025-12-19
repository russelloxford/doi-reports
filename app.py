"""
DOI (Division of Interest) Generator
A Streamlit application for generating Tract-Based Ownership and Unit-Based DOI reports.
"""

import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import warnings
warnings.filterwarnings('ignore')

# Page configuration
st.set_page_config(
    page_title="DOI Generator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for styling
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@400;600;700&family=Playfair+Display:wght@700&display=swap');
    
    .main-header {
        font-family: 'Playfair Display', serif;
        font-size: 2.5rem;
        font-weight: 700;
        color: #1a365d;
        margin-bottom: 0.5rem;
        text-align: center;
    }
    
    .sub-header {
        font-family: 'Source Sans Pro', sans-serif;
        font-size: 1.1rem;
        color: #4a5568;
        text-align: center;
        margin-bottom: 2rem;
    }
    
    .info-box {
        background-color: #ebf8ff;
        border-left: 4px solid #4299e1;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    
    .success-box {
        background-color: #f0fff4;
        border-left: 4px solid #48bb78;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    
    .warning-box {
        background-color: #fffaf0;
        border-left: 4px solid #ed8936;
        padding: 1rem;
        border-radius: 0 8px 8px 0;
        margin: 1rem 0;
    }
    
    div[data-testid="stMetricValue"] {
        font-size: 1.8rem;
        font-weight: 700;
        color: #2d3748;
    }
</style>
""", unsafe_allow_html=True)

# Helper functions
def safe_float(val, default=0):
    """Safely convert a value to float."""
    if val is None or pd.isna(val):
        return default
    if isinstance(val, str):
        val = val.strip()
        if val == '' or val == ' ':
            return default
        try:
            return float(val)
        except:
            return default
    try:
        return float(val)
    except:
        return default


def normalize_tract(val):
    """Normalize a tract value to a consistent string format."""
    if val is None or pd.isna(val):
        return ''
    val_str = str(val).strip()
    # If it's a numeric value like "1.0", convert to "1"
    try:
        num = float(val_str)
        if num == int(num):
            return str(int(num))
        return val_str
    except (ValueError, TypeError):
        return val_str


def tract_sort_key(tract):
    """Create a sort key for tract numbers that handles both numeric and text tracts.

    Sorts numerically for pure numbers, alphabetically for text tracts.
    Examples: 1, 2, 10, 11, Oram 1, Oram 2, Oram 10
    """
    import re
    tract_str = str(tract)

    # Try to parse as a pure number first
    try:
        num = float(tract_str)
        # Pure numbers sort first, using a tuple (0, number, '')
        return (0, num, '')
    except (ValueError, TypeError):
        pass

    # For text tracts, extract any trailing number for natural sorting
    # e.g., "Oram 10" -> ("Oram ", 10)
    match = re.match(r'^(.*?)(\d+)$', tract_str)
    if match:
        prefix = match.group(1)
        num = int(match.group(2))
        # Text tracts sort after numbers, using tuple (1, prefix, number)
        return (1, prefix.lower(), num)

    # Pure text with no numbers
    return (1, tract_str.lower(), 0)


def load_combined_data(uploaded_file):
    """Load and validate the Combined data file."""
    try:
        xl = pd.ExcelFile(uploaded_file)
        
        # Try to find the Combined sheet or use the data as-is
        if 'Combined' in xl.sheet_names:
            df = pd.read_excel(uploaded_file, sheet_name='Combined')
        else:
            # Try to find a sheet with the required columns
            for sheet in xl.sheet_names:
                df = pd.read_excel(uploaded_file, sheet_name=sheet)
                if 'OWNER' in df.columns and 'TYPE' in df.columns:
                    break
            else:
                # Check if first row is header
                df = pd.read_excel(uploaded_file, sheet_name=xl.sheet_names[0], header=None)
                # Find the header row
                for i, row in df.iterrows():
                    if 'OWNER' in str(row.values):
                        df = pd.read_excel(uploaded_file, sheet_name=xl.sheet_names[0], header=i)
                        break
        
        # Validate required columns
        required_cols = ['OWNER', 'TYPE', 'TRACT']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            return None, f"Missing required columns: {', '.join(missing)}"
        
        # Clean up TRACT column - normalize to consistent string format
        df['TRACT'] = df['TRACT'].apply(normalize_tract)
        df = df[df['TRACT'] != '']  # Remove empty tract values
        
        return df, None
        
    except Exception as e:
        return None, f"Error loading file: {str(e)}"


def load_tract_allocations(uploaded_file):
    """Load tract allocation data from Schedule file."""
    try:
        df = pd.read_excel(uploaded_file, sheet_name='Tract List', header=None)
        
        # Find the tract data section (look for "Tract" header)
        tract_start = None
        for i, row in df.iterrows():
            if str(row[0]).strip().lower() == 'tract':
                tract_start = i
                break
        
        if tract_start is None:
            return None, "Could not find Tract allocation data in file"
        
        # Read tract allocations
        allocations = {}
        for i in range(tract_start + 1, len(df)):
            row = df.iloc[i]
            tract = row[0]
            if pd.isna(tract) or str(tract).strip().upper() == 'TOTAL UNIT ACRES':
                break
            tract_key = normalize_tract(tract)
            if tract_key == '':
                continue
            allocation = safe_float(row[3], 0)  # Tract Allocation column
            acres = safe_float(row[2], 0)  # Acres column
            legal_desc = str(row[1]) if pd.notna(row[1]) else ''
            allocations[tract_key] = {
                'allocation': allocation,
                'acres': acres,
                'legal_description': legal_desc
            }
        
        if not allocations:
            return None, "No tract allocations found in file"
        
        return allocations, None
        
    except Exception as e:
        return None, f"Error loading tract allocations: {str(e)}"


def create_tract_based_workbook(df):
    """Create a Tract-Based Ownership Excel workbook (organized by tract)."""
    wb = Workbook()
    
    # Styles
    tnr_font = Font(name='Times New Roman', size=10)
    tnr_bold = Font(name='Times New Roman', size=10, bold=True)
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    tract_info_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='bottom')
    bottom_align = Alignment(vertical='bottom')
    wrap_align = Alignment(wrap_text=True, vertical='bottom', horizontal='left')
    narrow_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75)
    
    # Get tract info if available in the dataframe
    tract_info = {}
    for tract in df['TRACT'].unique():
        tract_data = df[df['TRACT'] == tract].iloc[0] if len(df[df['TRACT'] == tract]) > 0 else None
        if tract_data is not None:
            tract_info[tract] = {
                'legal_description': str(tract_data.get('Legal Description', '')) if pd.notna(tract_data.get('Legal Description')) else '',
                'gross_acres': safe_float(tract_data.get('Tract Gross Acres', tract_data.get('NET ACRES', 0)))
            }
    
    # ==================
    # TRACT LIST SHEET
    # ==================
    ws = wb.active
    ws.title = 'Tract List'
    
    # Create a simple tract list
    tract_headers = ['TRACT', 'LEGAL DESCRIPTION', 'GROSS ACRES']
    for col, header in enumerate(tract_headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = tnr_bold
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 15
    
    row_num = 2
    for tract in sorted(df['TRACT'].unique(), key=tract_sort_key):
        info = tract_info.get(tract, {'legal_description': '', 'gross_acres': 0})
        ws.cell(row=row_num, column=1, value=tract).font = tnr_font
        ws.cell(row=row_num, column=1).alignment = center_align
        ws.cell(row=row_num, column=1).border = thin_border
        ws.cell(row=row_num, column=2, value=info['legal_description']).font = tnr_font
        ws.cell(row=row_num, column=2).border = thin_border
        ws.cell(row=row_num, column=3, value=info['gross_acres']).font = tnr_font
        ws.cell(row=row_num, column=3).number_format = '0.00'
        ws.cell(row=row_num, column=3).border = thin_border
        row_num += 1
    
    ws.page_margins = narrow_margins
    ws.page_setup.orientation = 'landscape'
    ws.oddHeader.center.text = 'Tract List'
    ws.oddFooter.center.text = 'Page &P of &N'
    
    # ==================
    # INTEREST TYPE SHEETS
    # ==================
    interest_types = [
        ('MI', 'LORI', 'Landowner Royalty Interests'),
        ('NPRI', 'NPRI', 'Non-Participating Royalty Interests'),
        ('ORI', 'ORI', 'Overriding Royalty Interests'),
        ('WI', 'WI', 'Working Interests')
    ]

    # Build LORI lookup for WI sheet
    lori_lookup = {}
    mi_df = df[df['TYPE'] == 'MI']
    for tract in mi_df['TRACT'].unique():
        tract_mi = mi_df[mi_df['TRACT'] == tract]
        lori_lookup[tract] = {}
        for _, row in tract_mi.iterrows():
            lease = str(row.get('LEASE NO.', '') or '').strip()
            lori = safe_float(row.get('LEASE ROYALTY', 0))
            if lease:
                lori_lookup[tract][lease] = lori
            if 'default' not in lori_lookup[tract]:
                lori_lookup[tract]['default'] = lori

    for type_code, sheet_name, full_name in interest_types:
        type_df = df[df['TYPE'] == type_code].copy()
        
        # Filter out "None." owners
        type_df = type_df[~type_df['OWNER'].astype(str).str.lower().str.strip().isin(['none.', 'none', 'nan', ''])]
        if type_df.empty:
            continue
        
        ws = wb.create_sheet(sheet_name)
        
        # Define headers and column widths based on interest type
        if type_code == 'MI':
            headers = ['OWNER', 'TRACT', 'LEASE NO.', 'REQ', '', 'MI', 'x', 'LORI', '-', 'NPRI', '=', 'TRACT NRI', '', 'NET ACRES', '', 'BURDENED WI OWNER(S)']
            col_widths = [45, 12, 12, 12, 3, 12, 3, 12, 3, 12, 3, 12, 3, 12, 3, 50]
            nri_cols = [6, 8, 10, 12]
            acres_col = 14
            tract_nri_col = 12
        elif type_code == 'NPRI':
            headers = ['OWNER', 'TRACT', 'LEASE NO.', 'REQ', '', 'NPRI', 'x', 'INTEREST BURDENED', 'x', 'SHARE OF NPRI', '=', 'TRACT NRI', '', 'BURDENED MI OWNER']
            col_widths = [45, 12, 12, 12, 3, 12, 3, 15, 3, 14, 3, 12, 3, 50]
            nri_cols = [6, 8, 10, 12]
            acres_col = None
            tract_nri_col = 12
        elif type_code == 'ORI':
            headers = ['OWNER', 'TRACT', 'LEASE NO.', 'REQ', '', 'ORI', 'x', 'SHARE OF ORI', 'x', 'INTEREST BURDENED', '=', 'TRACT NRI', '', 'ACRES BURDENED', '', 'BURDENED WI OWNER(S)']
            col_widths = [45, 12, 12, 12, 3, 12, 3, 14, 3, 15, 3, 12, 3, 14, 3, 50]
            nri_cols = [6, 8, 10, 12]
            acres_col = 14
            tract_nri_col = 12
        else:  # WI
            headers = ['OWNER', 'TRACT', 'LEASE NO.', 'REQ', '', 'WI', '', 'NET ACRES', '', '1', '-', 'LORI', '-', 'ORI BURDENS', 'x', 'WI (TRACT)', '=', 'TRACT NRI']
            col_widths = [45, 12, 12, 12, 3, 12, 3, 12, 2, 5, 2, 12, 4, 12, 3, 12, 4, 12]
            nri_cols = [6, 12, 14, 16, 18]
            acres_col = 8
            tract_nri_col = 18
        
        # Set column widths
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = tnr_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            cell.border = Border(bottom=thin_side)
        
        # Process data BY TRACT
        current_row = 2

        for tract in sorted(type_df['TRACT'].unique(), key=tract_sort_key):
            tract_data = type_df[type_df['TRACT'] == tract].sort_values('OWNER')
            info = tract_info.get(tract, {'legal_description': '', 'gross_acres': 0})
            
            # Empty row before tract
            current_row += 1
            
            # Tract info box (3 rows with gray fill)
            # Row 1: Tract No.
            ws.cell(row=current_row, column=1, value='Tract No.:')
            ws.cell(row=current_row, column=1).font = tnr_font
            ws.cell(row=current_row, column=1).fill = tract_info_fill
            ws.cell(row=current_row, column=2, value=tract)
            ws.cell(row=current_row, column=2).font = tnr_font
            ws.cell(row=current_row, column=2).fill = tract_info_fill
            for col in range(3, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = tract_info_fill
            current_row += 1
            
            # Row 2: Gross Acres
            ws.cell(row=current_row, column=1, value='Gross Acres:')
            ws.cell(row=current_row, column=1).font = tnr_font
            ws.cell(row=current_row, column=1).fill = tract_info_fill
            ws.cell(row=current_row, column=2, value=info['gross_acres'])
            ws.cell(row=current_row, column=2).font = tnr_font
            ws.cell(row=current_row, column=2).fill = tract_info_fill
            ws.cell(row=current_row, column=2).number_format = '0.00'
            for col in range(3, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = tract_info_fill
            current_row += 1
            
            # Row 3: Legal Description
            ws.cell(row=current_row, column=1, value='Legal Description:')
            ws.cell(row=current_row, column=1).font = tnr_font
            ws.cell(row=current_row, column=1).fill = tract_info_fill
            ws.cell(row=current_row, column=2, value=info['legal_description'])
            ws.cell(row=current_row, column=2).font = tnr_font
            ws.cell(row=current_row, column=2).fill = tract_info_fill
            ws.cell(row=current_row, column=2).alignment = wrap_align
            for col in range(3, len(headers) + 1):
                ws.cell(row=current_row, column=col).fill = tract_info_fill
            current_row += 1
            
            # Empty row after tract info
            current_row += 1
            
            # Data rows for this tract
            tract_nri_total = 0
            mi_total = 0
            
            for _, row in tract_data.iterrows():
                owner = str(row.get('OWNER', ''))
                tract_nri = safe_float(row.get('TRACT NRI', 0))
                tract_nri_total += tract_nri
                
                lease_no = str(row.get('LEASE NO.', '') or '').strip()
                if lease_no == 'nan':
                    lease_no = ''
                req = str(row.get('REQ', '') or '').strip()
                if req == 'nan':
                    req = ''
                
                if type_code == 'MI':
                    mi_val = safe_float(row.get('DECIMAL INTEREST', 0))
                    mi_total += mi_val
                    burdened_wi = str(row.get('Burdened WI Owners', '')) if pd.notna(row.get('Burdened WI Owners')) else ''
                    values = [
                        owner, tract, lease_no, req, '',
                        mi_val, 'x',
                        safe_float(row.get('LEASE ROYALTY', 0)), '-',
                        safe_float(row.get('NPRI BURDENS', 0)), '=',
                        tract_nri, '',
                        safe_float(row.get('NET ACRES', 0)), '',
                        burdened_wi
                    ]
                elif type_code == 'NPRI':
                    burdened_mi = str(row.get('Burdened WI Owners', '')) if pd.notna(row.get('Burdened WI Owners')) else ''
                    values = [
                        owner, tract, lease_no, req, '',
                        safe_float(row.get('NPRI', row.get('DECIMAL INTEREST', 0))), 'x',
                        safe_float(row.get('INTEREST BURDENED', 0)), 'x',
                        safe_float(row.get('SHARE OF NPRI', 0)), '=',
                        tract_nri, '',
                        burdened_mi
                    ]
                elif type_code == 'ORI':
                    burdened_wi = str(row.get('Burdened WI Owners', '')) if pd.notna(row.get('Burdened WI Owners')) else ''
                    values = [
                        owner, tract, lease_no, req, '',
                        safe_float(row.get('ORI', row.get('DECIMAL INTEREST', 0))), 'x',
                        safe_float(row.get('SHARE OF ORI', 0)), 'x',
                        safe_float(row.get('INTEREST BURDENED', 0)), '=',
                        tract_nri, '',
                        safe_float(row.get('ACRES BURDENED', 0)), '',
                        burdened_wi
                    ]
                else:  # WI
                    wi_val = safe_float(row.get('DECIMAL INTEREST', 0))
                    # Get LORI for this tract/lease
                    tract_lori = lori_lookup.get(tract, {})
                    lori_val = tract_lori.get(lease_no, tract_lori.get('default', 0))
                    ori_burdens = safe_float(row.get('ORI BURDENS', 0))
                    wi_tract = safe_float(row.get('WI (TRACT)', wi_val))
                    
                    values = [
                        owner, tract, lease_no, req, '',
                        wi_val, '',
                        safe_float(row.get('NET ACRES', 0)), '',
                        1, '-',
                        lori_val, '-',
                        ori_burdens, 'x',
                        wi_tract, '=',
                        tract_nri
                    ]
                
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.font = tnr_font
                    cell.alignment = bottom_align
                    
                    if col in nri_cols:
                        cell.number_format = '0.00000000'
                    elif acres_col and col == acres_col:
                        cell.number_format = '0.000000'
                    
                    if col in [2, 3, 4, 5, 7, 9, 10, 11, 13, 15, 17]:
                        cell.alignment = center_align
                
                current_row += 1
            
            # TOTALS row
            ws.cell(row=current_row, column=1, value='TOTALS')
            ws.cell(row=current_row, column=1).font = tnr_bold
            
            if type_code == 'MI':
                ws.cell(row=current_row, column=6, value=mi_total)
                ws.cell(row=current_row, column=6).font = tnr_bold
                ws.cell(row=current_row, column=6).number_format = '0.00000000'
            
            ws.cell(row=current_row, column=tract_nri_col, value=tract_nri_total)
            ws.cell(row=current_row, column=tract_nri_col).font = tnr_bold
            ws.cell(row=current_row, column=tract_nri_col).number_format = '0.00000000'
            
            current_row += 1
        
        # Page setup
        ws.page_margins = narrow_margins
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = '1:1'
        ws.oddHeader.center.text = f'Tract-Based {full_name}'
        ws.oddFooter.center.text = 'Page &P of &N'
    
    # ==================
    # UNIT RECAP SHEET
    # ==================
    ws = wb.create_sheet('Unit Recap')
    
    headers = ['TRACT', 'LORI NRI', 'NPRI NRI', 'ORI NRI', 'WI NRI', 'TOTAL NRI']
    col_widths = [15, 14, 14, 14, 14, 14]
    
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = tnr_bold
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row = 2
    totals = {'MI': 0, 'NPRI': 0, 'ORI': 0, 'WI': 0}

    for tract in sorted(df['TRACT'].unique(), key=tract_sort_key):
        lori_nri = sum(safe_float(r.get('TRACT NRI', 0)) for _, r in df[(df['TYPE'] == 'MI') & (df['TRACT'] == tract)].iterrows())
        npri_nri = sum(safe_float(r.get('TRACT NRI', 0)) for _, r in df[(df['TYPE'] == 'NPRI') & (df['TRACT'] == tract) & (~df['OWNER'].astype(str).str.lower().str.strip().isin(['none.', 'none']))].iterrows())
        ori_nri = sum(safe_float(r.get('TRACT NRI', 0)) for _, r in df[(df['TYPE'] == 'ORI') & (df['TRACT'] == tract)].iterrows())
        wi_nri = sum(safe_float(r.get('TRACT NRI', 0)) for _, r in df[(df['TYPE'] == 'WI') & (df['TRACT'] == tract)].iterrows())
        
        totals['MI'] += lori_nri
        totals['NPRI'] += npri_nri
        totals['ORI'] += ori_nri
        totals['WI'] += wi_nri
        
        total_nri = lori_nri + npri_nri + ori_nri + wi_nri
        
        values = [tract, lori_nri, npri_nri, ori_nri, wi_nri, total_nri]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = tnr_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = center_align
            else:
                cell.number_format = '0.00000000'
        
        current_row += 1
    
    # Total row
    total_all = totals['MI'] + totals['NPRI'] + totals['ORI'] + totals['WI']
    values = ['TOTAL', totals['MI'], totals['NPRI'], totals['ORI'], totals['WI'], total_all]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = tnr_bold
        cell.border = thin_border
        if col == 1:
            cell.alignment = center_align
        else:
            cell.number_format = '0.00000000'
    
    ws.page_margins = narrow_margins
    ws.page_setup.orientation = 'landscape'
    ws.oddHeader.center.text = 'Unit Recap'
    ws.oddFooter.center.text = 'Page &P of &N'
    
    return wb


def create_unit_based_workbook(df, allocations, schedule_file):
    """Create a Unit-Based DOI Excel workbook (organized by owner)."""
    wb = Workbook()

    # Copy the Tract List sheet from the schedule file as the first sheet
    from openpyxl import load_workbook
    schedule_file.seek(0)  # Reset file pointer
    source_wb = load_workbook(schedule_file)
    if 'Tract List' in source_wb.sheetnames:
        source_ws = source_wb['Tract List']
        # Rename the default sheet to 'Tract List'
        ws = wb.active
        ws.title = 'Tract List'
        # Copy all cells from source to destination
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.border = cell.border.copy()
                    new_cell.fill = cell.fill.copy()
                    new_cell.number_format = cell.number_format
                    new_cell.protection = cell.protection.copy()
                    new_cell.alignment = cell.alignment.copy()
        # Copy column widths
        for col_letter, col_dim in source_ws.column_dimensions.items():
            ws.column_dimensions[col_letter].width = col_dim.width
        # Copy row heights
        for row_num, row_dim in source_ws.row_dimensions.items():
            ws.row_dimensions[row_num].height = row_dim.height
        # Copy merged cells
        for merged_range in source_ws.merged_cells.ranges:
            ws.merge_cells(str(merged_range))
    
    # Styles
    tnr_font = Font(name='Times New Roman', size=10)
    tnr_bold = Font(name='Times New Roman', size=10, bold=True)
    thin_side = Side(style='thin')
    medium_side = Side(style='medium')
    no_side = Side(style=None)
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    thin_border_no_bottom = Border(left=thin_side, right=thin_side, top=thin_side)
    header_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    owner_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='bottom')
    bottom_align = Alignment(vertical='bottom')
    wrap_align = Alignment(wrap_text=True, vertical='bottom', horizontal='left')
    narrow_margins = PageMargins(left=0.25, right=0.25, top=0.75, bottom=0.75)
    
    # Filter to only tracts in allocations
    valid_tracts = list(allocations.keys())
    df = df[df['TRACT'].isin(valid_tracts)].copy()
    
    # Build LORI lookup for WI sheet
    lori_lookup = {}
    mi_df = df[df['TYPE'] == 'MI']
    for tract in mi_df['TRACT'].unique():
        tract_mi = mi_df[mi_df['TRACT'] == tract]
        lori_lookup[tract] = {}
        for _, row in tract_mi.iterrows():
            lease = str(row.get('LEASE NO.', '') or '').strip()
            lori = safe_float(row.get('LEASE ROYALTY', 0))
            if lease:
                lori_lookup[tract][lease] = lori
            if 'default' not in lori_lookup[tract]:
                lori_lookup[tract]['default'] = lori

    interest_types = [
        ('MI', 'LORI', 'Landowner Royalty Interests'),
        ('NPRI', 'NPRI', 'Non-Participating Royalty Interests'),
        ('ORI', 'ORI', 'Overriding Royalty Interests'),
        ('WI', 'WI', 'Working Interests')
    ]
    
    for type_code, sheet_name, full_name in interest_types:
        type_df = df[df['TYPE'] == type_code].copy()

        # Filter out "None." owners
        type_df = type_df[~type_df['OWNER'].astype(str).str.lower().str.strip().isin(['none.', 'none', 'nan', ''])]

        if type_df.empty:
            continue

        ws = wb.create_sheet(sheet_name)
        
        # Define headers - include UNIT NRI column
        if type_code == 'MI':
            headers = ['TRACT', 'LEASE NO.', 'REQ', '', 'MI', 'x', 'LORI', '-', 'NPRI', '=', 'TRACT NRI', '', 'NET ACRES', '', 'UNIT NRI']
            col_widths = [13, 12, 13, 3, 12, 3, 12, 3, 12, 3, 12, 3, 12, 3, 14]
            nri_cols = [5, 7, 9, 11, 15]
            acres_col = 13
        elif type_code == 'NPRI':
            headers = ['TRACT', 'LEASE NO.', 'REQ', '', 'NPRI', 'x', 'INTEREST BURDENED', 'x', 'SHARE OF NPRI', '=', 'TRACT NRI', '', 'UNIT NRI']
            col_widths = [13, 12, 10, 3, 12, 3, 15, 3, 14, 3, 12, 3, 13]
            nri_cols = [5, 7, 9, 11, 13]
            acres_col = None
        elif type_code == 'ORI':
            headers = ['TRACT', 'LEASE NO.', 'REQ', '', 'ORI', 'x', 'SHARE OF ORI', 'x', 'INTEREST BURDENED', '=', 'TRACT NRI', '', 'ACRES BURDENED', '', 'UNIT NRI']
            col_widths = [14, 12, 10, 3, 12, 3, 11, 3, 12, 3, 11, 3, 11, 3, 13]
            nri_cols = [5, 7, 9, 11, 15]
            acres_col = 13
        else:  # WI
            headers = ['TRACT', 'LEASE NO.', 'REQ', '', 'WI', '', 'NET ACRES', '', '1', '-', 'LORI', '-', 'ORI BURDENS', 'x', 'WI (TRACT)', '=', 'TRACT NRI', '', 'UNIT NRI']
            col_widths = [14, 12, 13, 3, 12, 3, 12, 2, 5, 2, 12, 4, 12, 3, 12, 4, 12, 2, 12]
            nri_cols = [5, 11, 13, 15, 17, 19]
            acres_col = 7
        
        # Set column widths
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = tnr_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            cell.border = Border(bottom=thin_side)
        
        # Group by owner
        type_df = type_df.sort_values(['OWNER', 'TRACT'])
        current_row = 2
        
        for owner in type_df['OWNER'].unique():
            if pd.isna(owner) or str(owner).strip().lower() in ['none.', 'none', '']:
                continue
                
            owner_data = type_df[type_df['OWNER'] == owner].sort_values('TRACT')
            
            # Blank row before owner
            current_row += 1
            
            # Owner name row
            ws.cell(row=current_row, column=1, value='Owner Name:')
            ws.cell(row=current_row, column=1).font = tnr_bold
            ws.cell(row=current_row, column=1).fill = owner_fill
            ws.cell(row=current_row, column=1).alignment = bottom_align

            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(headers))
            ws.cell(row=current_row, column=2, value=str(owner))
            ws.cell(row=current_row, column=2).font = tnr_font
            ws.cell(row=current_row, column=2).fill = owner_fill
            ws.cell(row=current_row, column=2).alignment = wrap_align

            # Apply border with both top and bottom medium borders for single-row owner box
            for col in range(1, len(headers) + 1):
                left = medium_side if col == 1 else no_side
                right = medium_side if col == len(headers) else no_side
                ws.cell(row=current_row, column=col).border = Border(left=left, right=right, top=medium_side, bottom=medium_side)
                ws.cell(row=current_row, column=col).fill = owner_fill

            current_row += 1
            current_row += 1  # Blank row after owner info
            
            # Data rows
            for idx, (_, row) in enumerate(owner_data.iterrows()):
                tract = row['TRACT']
                allocation = allocations.get(tract, {}).get('allocation', 0)
                tract_nri = safe_float(row.get('TRACT NRI', 0))
                unit_nri = tract_nri * allocation
                
                lease_no = str(row.get('LEASE NO.', '') or '').strip()
                if lease_no == 'nan':
                    lease_no = ''
                req = str(row.get('REQ', '') or '').strip()
                if req == 'nan':
                    req = ''
                
                if type_code == 'MI':
                    values = [
                        tract, lease_no, req, '',
                        safe_float(row.get('DECIMAL INTEREST', 0)), 'x',
                        safe_float(row.get('LEASE ROYALTY', 0)), '-',
                        safe_float(row.get('NPRI BURDENS', 0)), '=',
                        tract_nri, '',
                        safe_float(row.get('NET ACRES', 0)), '',
                        unit_nri
                    ]
                elif type_code == 'NPRI':
                    values = [
                        tract, lease_no, req, '',
                        safe_float(row.get('NPRI', row.get('DECIMAL INTEREST', 0))), 'x',
                        safe_float(row.get('INTEREST BURDENED', 0)), 'x',
                        safe_float(row.get('SHARE OF NPRI', 0)), '=',
                        tract_nri, '',
                        unit_nri
                    ]
                elif type_code == 'ORI':
                    values = [
                        tract, lease_no, req, '',
                        safe_float(row.get('ORI', row.get('DECIMAL INTEREST', 0))), 'x',
                        safe_float(row.get('SHARE OF ORI', 0)), 'x',
                        safe_float(row.get('INTEREST BURDENED', 0)), '=',
                        tract_nri, '',
                        safe_float(row.get('ACRES BURDENED', 0)), '',
                        unit_nri
                    ]
                else:  # WI
                    wi_val = safe_float(row.get('DECIMAL INTEREST', 0))
                    tract_lori = lori_lookup.get(tract, {})
                    lori_val = tract_lori.get(lease_no, tract_lori.get('default', 0))
                    ori_burdens = safe_float(row.get('ORI BURDENS', 0))
                    wi_tract = safe_float(row.get('WI (TRACT)', wi_val))
                    
                    values = [
                        tract, lease_no, req, '',
                        wi_val, '',
                        safe_float(row.get('NET ACRES', 0)), '',
                        1, '-',
                        lori_val, '-',
                        ori_burdens, 'x',
                        wi_tract, '=',
                        tract_nri, '',
                        unit_nri
                    ]
                
                is_last = idx == len(owner_data) - 1
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.font = tnr_font
                    cell.border = thin_border if is_last else thin_border_no_bottom
                    
                    if col in nri_cols:
                        cell.number_format = '0.00000000'
                    elif acres_col and col == acres_col:
                        cell.number_format = '0.000000'
                    
                    if col in [1, 2, 3, 4, 6, 8, 9, 10, 12, 14, 16, 18]:
                        cell.alignment = center_align
                
                current_row += 1
            
            # TOTAL row
            ws.cell(row=current_row, column=1, value='TOTAL')
            ws.cell(row=current_row, column=1).font = tnr_bold
            ws.cell(row=current_row, column=1).border = thin_border
            
            unit_nri_total = sum(
                safe_float(r.get('TRACT NRI', 0)) * allocations.get(r['TRACT'], {}).get('allocation', 0)
                for _, r in owner_data.iterrows()
            )
            
            for col in range(2, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.border = thin_border
                cell.font = tnr_bold
            
            # Unit NRI total in last column
            ws.cell(row=current_row, column=len(headers), value=unit_nri_total)
            ws.cell(row=current_row, column=len(headers)).number_format = '0.00000000'
            ws.cell(row=current_row, column=len(headers)).font = tnr_bold
            
            current_row += 1
        
        # Page setup
        ws.page_margins = narrow_margins
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_title_rows = '1:1'
        ws.oddHeader.center.text = f'Unit-Based {full_name}'
        ws.oddFooter.center.text = 'Page &P of &N'
    
    # ==================
    # UNIT RECAP SHEET
    # ==================
    ws = wb.create_sheet('Unit Recap')
    
    headers = ['TRACT', 'LORI NRI', 'NPRI NRI', 'ORI NRI', 'WI NRI', 'TOTAL NRI']
    col_widths = [15, 14, 14, 14, 14, 14]
    
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = tnr_bold
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    current_row = 2
    totals = {'MI': 0, 'NPRI': 0, 'ORI': 0, 'WI': 0}

    for tract in sorted(allocations.keys(), key=tract_sort_key):
        allocation = allocations[tract]['allocation']
        
        lori_nri = sum(safe_float(r.get('TRACT NRI', 0)) * allocation for _, r in df[(df['TYPE'] == 'MI') & (df['TRACT'] == tract)].iterrows())
        npri_nri = sum(safe_float(r.get('TRACT NRI', 0)) * allocation for _, r in df[(df['TYPE'] == 'NPRI') & (df['TRACT'] == tract) & (~df['OWNER'].astype(str).str.lower().str.strip().isin(['none.', 'none']))].iterrows())
        ori_nri = sum(safe_float(r.get('TRACT NRI', 0)) * allocation for _, r in df[(df['TYPE'] == 'ORI') & (df['TRACT'] == tract)].iterrows())
        wi_nri = sum(safe_float(r.get('TRACT NRI', 0)) * allocation for _, r in df[(df['TYPE'] == 'WI') & (df['TRACT'] == tract)].iterrows())
        
        totals['MI'] += lori_nri
        totals['NPRI'] += npri_nri
        totals['ORI'] += ori_nri
        totals['WI'] += wi_nri
        
        total_nri = lori_nri + npri_nri + ori_nri + wi_nri
        
        values = [tract, lori_nri, npri_nri, ori_nri, wi_nri, total_nri]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=col, value=val)
            cell.font = tnr_font
            cell.border = thin_border
            if col == 1:
                cell.alignment = center_align
            else:
                cell.number_format = '0.00000000'
        
        current_row += 1
    
    # Total row
    total_all = totals['MI'] + totals['NPRI'] + totals['ORI'] + totals['WI']
    values = ['UNIT NRI TOTAL', totals['MI'], totals['NPRI'], totals['ORI'], totals['WI'], total_all]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=current_row, column=col, value=val)
        cell.font = tnr_bold
        cell.border = thin_border
        if col == 1:
            cell.alignment = center_align
        else:
            cell.number_format = '0.00000000'
    
    ws.page_margins = narrow_margins
    ws.page_setup.orientation = 'landscape'
    ws.oddHeader.center.text = 'Unit Recap'
    ws.oddFooter.center.text = 'Page &P of &N'
    
    return wb, total_all


def to_excel_bytes(workbook):
    """Convert workbook to bytes for download."""
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


# Main app
def main():
    st.markdown('<h1 class="main-header">📊 DOI Generator</h1>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Generate Tract-Based Ownership or Unit-Based Division of Interest Reports</p>', unsafe_allow_html=True)
    
    # Sidebar for report type selection
    with st.sidebar:
        st.markdown("### Report Configuration")
        report_type = st.radio(
            "Select Report Type:",
            ["Tract-Based Ownership", "Unit-Based DOI"],
            help="Tract-Based organizes data by tract. Unit-Based organizes by owner and applies allocation factors."
        )
        
        st.markdown("---")
        st.markdown("### Required Files")
        
        if report_type == "Tract-Based Ownership":
            st.markdown("""
            <div class="info-box">
            <strong>Combined Data File</strong><br>
            Excel file with ownership data including:<br>
            • OWNER, TYPE, TRACT columns<br>
            • TRACT NRI values<br>
            • Interest-specific fields
            </div>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <div class="info-box">
            <strong>1. Combined Data File</strong><br>
            Same as Tract-Based requirements
            </div>
            """, unsafe_allow_html=True)
            st.markdown("""
            <div class="info-box">
            <strong>2. Schedule File</strong><br>
            Excel file with Tract List sheet containing:<br>
            • Tract numbers<br>
            • Tract Allocation factors
            </div>
            """, unsafe_allow_html=True)
    
    # Main content area
    st.markdown("---")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown(f"### {report_type}")
        
        # File uploads
        combined_file = st.file_uploader(
            "Upload Combined Data File (Excel)",
            type=['xlsx', 'xls'],
            help="Excel file containing ownership data with OWNER, TYPE, TRACT, and TRACT NRI columns"
        )
        
        schedule_file = None
        if report_type == "Unit-Based DOI":
            schedule_file = st.file_uploader(
                "Upload Schedule File (Excel)",
                type=['xlsx', 'xls'],
                help="Excel file with Tract List containing tract allocations"
            )
    
    with col2:
        st.markdown("### Preview")
        
        if combined_file:
            df, error = load_combined_data(combined_file)
            if error:
                st.error(error)
            else:
                st.metric("Total Records", len(df))
                st.metric("Unique Tracts", df['TRACT'].nunique())
                st.metric("Unique Owners", df['OWNER'].nunique())
                
                type_counts = df['TYPE'].value_counts()
                st.markdown("**By Interest Type:**")
                for t, count in type_counts.items():
                    if pd.notna(t):
                        st.text(f"  {t}: {count}")
    
    st.markdown("---")
    
    # Generate button and output
    if combined_file:
        df, error = load_combined_data(combined_file)
        
        if error:
            st.error(error)
        else:
            allocations = None
            alloc_error = None
            
            if report_type == "Unit-Based DOI":
                if schedule_file:
                    allocations, alloc_error = load_tract_allocations(schedule_file)
                    if alloc_error:
                        st.error(alloc_error)
                    else:
                        st.success(f"✅ Loaded {len(allocations)} tract allocations")
                        
                        # Show allocation preview
                        with st.expander("View Tract Allocations"):
                            alloc_df = pd.DataFrame([
                                {'Tract': k, 'Allocation': v['allocation'], 'Acres': v['acres']}
                                for k, v in sorted(allocations.items(), key=lambda x: tract_sort_key(x[0]))
                            ])
                            st.dataframe(alloc_df, use_container_width=True)
                else:
                    st.warning("⚠️ Please upload a Schedule file with tract allocations")
            
            # Generate button
            can_generate = (report_type == "Tract-Based Ownership") or (report_type == "Unit-Based DOI" and allocations is not None)
            
            if can_generate:
                if st.button("🚀 Generate Report", type="primary", use_container_width=True):
                    with st.spinner("Generating report..."):
                        try:
                            if report_type == "Tract-Based Ownership":
                                wb = create_tract_based_workbook(df)
                                filename = "Tract_Based_Ownership.xlsx"
                                
                                st.markdown("""
                                <div class="success-box">
                                <strong>✅ Report Generated Successfully!</strong><br>
                                Your Tract-Based Ownership report is ready for download.
                                </div>
                                """, unsafe_allow_html=True)
                                
                            else:  # Unit-Based DOI
                                wb, total_nri = create_unit_based_workbook(df, allocations, schedule_file)
                                filename = "Unit_Based_DOI.xlsx"
                                
                                # Show validation
                                if abs(total_nri - 1.0) < 0.0001:
                                    st.markdown(f"""
                                    <div class="success-box">
                                    <strong>✅ Report Generated Successfully!</strong><br>
                                    Unit NRI Total: {total_nri:.8f} ✓ (Verified = 1.00000000)
                                    </div>
                                    """, unsafe_allow_html=True)
                                else:
                                    st.markdown(f"""
                                    <div class="warning-box">
                                    <strong>⚠️ Report Generated with Warning</strong><br>
                                    Unit NRI Total: {total_nri:.8f}<br>
                                    Expected: 1.00000000<br>
                                    Please verify source data.
                                    </div>
                                    """, unsafe_allow_html=True)
                            
                            # Download button
                            excel_bytes = to_excel_bytes(wb)
                            st.download_button(
                                label=f"📥 Download {filename}",
                                data=excel_bytes,
                                file_name=filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            
                        except Exception as e:
                            st.error(f"Error generating report: {str(e)}")
    
    # Footer
    st.markdown("---")
    st.markdown("""
    <div style="text-align: center; color: #718096; font-size: 0.9rem;">
    DOI Generator | Generates professional Division of Interest reports from ownership data
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
